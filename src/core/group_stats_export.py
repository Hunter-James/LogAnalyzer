"""Экспорт результата GroupStatsWorker в файлы: HTML/XLSX/CSV/JSON.

Все экспортёры принимают batches: {batch_id: bucket} (см. _new_batch_bucket
в group_stats_worker.py). Сортировка как в дереве диалога: сначала
хронологически по first_ts/first_file, потом NO_BATCH («вне партии»)
в конец.

Зависимости:
- HTML / CSV / JSON — только stdlib.
- XLSX — openpyxl (опционально, проверка в _try_import_openpyxl)."""

import csv
import json
import os
import html
from datetime import datetime

from core.models import NO_BATCH


# Метки счётчиков (один источник истины для всех экспортёров)
_COUNTER_LABELS = [
    ('printed', 'Напечатано'),
    ('scanned', 'Прочитано'),
    ('noread', 'No read'),
    ('verified', 'Верифицировано'),
    ('rejected', 'Отбраковано'),
    ('not_verified', 'Не верифицировано'),
]


def _sort_key(item):
    """Та же сортировка что в GroupStatsDialog._populate_tree: NO_BATCH в
    конец, остальные хронологически по first_ts."""
    bid, data = item
    no_batch_flag = 1 if bid == NO_BATCH else 0
    ts = data.get('first_ts') or '99:99:99.999'
    first_file = data.get('first_file', '')
    return (no_batch_flag, first_file, ts)


def _batch_title(bid, data):
    """Заголовок партии для HTML/CSV — единая формулировка."""
    counters = data['counters']
    files = data['files']
    codes = data['codes']
    total_events = sum(counters.values())
    suffix = (f"{total_events:,} событий, {len(files)} файл(ов), "
              f"{len(codes):,} уникальных кодов")
    if bid == NO_BATCH:
        return f"Вне партии — {suffix}"
    return f"Партия {bid} — {suffix}"


def _try_import_openpyxl():
    """Возвращает модуль openpyxl или None. Не падаем при отсутствии —
    UI диалога экспорта сам скажет «установите openpyxl»."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        return None


# ===========================================================================
# CSV
# ===========================================================================

def export_csv(path, batches):
    """Плоский CSV: одна строка на партию + дополнительный sheet нет
    (CSV — один файл). Если нужны файлы/коды per-partition детально —
    делаем второй CSV рядом (path.replace('.csv', '_files.csv')).

    Возвращает (main_path, extra_paths) для UI-feedback'а."""
    extra_paths = []

    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f, delimiter=';')
        header = ['Партия']
        header += [label for _, label in _COUNTER_LABELS]
        header += ['Всего событий', 'Файлов', 'Уникальных кодов',
                   'Первое время', 'Первый файл',
                   'Последнее время', 'Последний файл']
        w.writerow(header)

        for bid, data in sorted(batches.items(), key=_sort_key):
            counters = data['counters']
            total_events = sum(counters.values())
            row = ['Вне партии' if bid == NO_BATCH else str(bid)]
            row += [counters[k] for k, _ in _COUNTER_LABELS]
            row += [
                total_events,
                len(data['files']),
                len(data['codes']),
                data.get('first_ts', ''),
                os.path.basename(data.get('first_file', '')),
                data.get('last_ts', ''),
                os.path.basename(data.get('last_file', '')),
            ]
            w.writerow(row)

    # Второй CSV: партия × файл, со счётчиком событий в каждом файле.
    files_path = path.replace('.csv', '_files.csv')
    if files_path == path:
        files_path = path + '_files.csv'
    with open(files_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['Партия', 'Файл', 'Событий'])
        for bid, data in sorted(batches.items(), key=_sort_key):
            label = 'Вне партии' if bid == NO_BATCH else str(bid)
            for fp in sorted(data['files'].keys()):
                w.writerow([label, os.path.basename(fp), data['files'][fp]])
    extra_paths.append(files_path)

    # Третий CSV: партия × код (только уникальные).
    codes_path = path.replace('.csv', '_codes.csv')
    if codes_path == path:
        codes_path = path + '_codes.csv'
    with open(codes_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['Партия', 'Код'])
        for bid, data in sorted(batches.items(), key=_sort_key):
            label = 'Вне партии' if bid == NO_BATCH else str(bid)
            for c in sorted(data['codes']):
                w.writerow([label, c])
    extra_paths.append(codes_path)

    return path, extra_paths


# ===========================================================================
# JSON
# ===========================================================================

def export_json(path, batches):
    """JSON со всеми данными. set кодов конвертируется в sorted list для
    воспроизводимости."""
    out = {
        'exported_at': datetime.now().isoformat(timespec='seconds'),
        'batches_count': len(batches),
        'batches': [],
    }
    for bid, data in sorted(batches.items(), key=_sort_key):
        out['batches'].append({
            'batch_id': bid if bid != NO_BATCH else None,
            'counters': dict(data['counters']),
            'files': [
                {'path': fp, 'events': cnt}
                for fp, cnt in sorted(data['files'].items())
            ],
            'first_ts': data.get('first_ts', ''),
            'first_file': data.get('first_file', ''),
            'last_ts': data.get('last_ts', ''),
            'last_file': data.get('last_file', ''),
            'codes': sorted(data['codes']),
        })
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    return path, []


# ===========================================================================
# HTML с <details>/<summary> — раскрывается в браузере
# ===========================================================================

_HTML_HEADER = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>Сводка партий по группе</title>
<style>
  body { font-family: 'Segoe UI', Arial, sans-serif; background: #1E1E1E; color: #E0E0E0;
         margin: 20px; }
  h1 { color: #FFFFFF; }
  .meta { color: #888; font-size: 0.9em; margin-bottom: 16px; }
  details { background: #2A2A2A; border: 1px solid #444; border-radius: 4px;
            padding: 8px 12px; margin: 6px 0; }
  details > summary { cursor: pointer; font-weight: bold; color: #4FB7E6;
                      padding: 4px 0; }
  details > summary:hover { color: #6BC7F0; }
  details details { background: #333; margin: 4px 0; }
  details details > summary { color: #C0C0C0; font-weight: normal; }
  table { border-collapse: collapse; margin: 4px 0 4px 8px; }
  table td, table th { border: 1px solid #444; padding: 3px 10px;
                       font-family: Consolas, monospace; font-size: 0.9em; }
  table th { background: #353535; text-align: left; }
  .info { color: #2E8B57; } .warn { color: #FFA500; } .error { color: #CD5C5C; }
  .muted { color: #888; }
  ul { margin: 4px 0 4px 24px; padding: 0; }
  ul li { font-family: Consolas, monospace; font-size: 0.9em; line-height: 1.4; }
  .toolbar { margin: 8px 0 16px; }
  .toolbar button { background: #2A82DA; color: white; border: none;
                    padding: 6px 14px; border-radius: 3px; cursor: pointer;
                    margin-right: 8px; }
  .toolbar button:hover { background: #3A92EA; }
</style>
<script>
function expandAll() {
  document.querySelectorAll('details').forEach(d => d.open = true);
}
function collapseAll() {
  document.querySelectorAll('details').forEach(d => d.open = false);
}
</script>
</head>
<body>
<h1>Сводка партий по группе</h1>
"""


def _counter_class(key, value):
    """Цветовая категория счётчика для HTML (info/warn/error/muted)."""
    if value == 0:
        return 'muted'
    if key in ('noread', 'not_verified'):
        return 'warn'
    if key == 'rejected':
        return 'error'
    return 'info'


def export_html(path, batches):
    """HTML с раскрывающимися <details>/<summary>. Кнопки «Раскрыть всё /
    Свернуть всё» вверху для удобной навигации."""
    parts = [_HTML_HEADER]
    parts.append(f'<div class="meta">Экспорт: '
                 f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, '
                 f'партий: {len(batches)}</div>')
    parts.append('<div class="toolbar">'
                 '<button onclick="expandAll()">Раскрыть всё</button>'
                 '<button onclick="collapseAll()">Свернуть всё</button>'
                 '</div>')

    for bid, data in sorted(batches.items(), key=_sort_key):
        counters = data['counters']
        files = data['files']
        codes = data['codes']
        title = html.escape(_batch_title(bid, data))
        parts.append(f'<details><summary>{title}</summary>')

        # Статистика
        parts.append('<details open><summary>📊 Статистика</summary><table>')
        parts.append('<tr><th>Категория</th><th>Значение</th></tr>')
        for key, label in _COUNTER_LABELS:
            n = counters[key]
            cls = _counter_class(key, n)
            parts.append(f'<tr><td>{html.escape(label)}</td>'
                         f'<td class="{cls}">{n:,}</td></tr>')
        parts.append('</table></details>')

        # Файлы
        if files:
            parts.append(f'<details><summary>📂 Файлы ({len(files)})</summary><table>')
            parts.append('<tr><th>Файл</th><th>Событий</th></tr>')
            for fp in sorted(files.keys()):
                parts.append(f'<tr><td>{html.escape(os.path.basename(fp))}</td>'
                             f'<td>{files[fp]:,}</td></tr>')
            parts.append('</table></details>')

        # Диапазон
        if data.get('first_ts'):
            parts.append('<details><summary>📅 Диапазон</summary><table>')
            parts.append(f'<tr><td>начало</td><td>{html.escape(data["first_ts"])}</td>'
                         f'<td>{html.escape(os.path.basename(data["first_file"]))}</td></tr>')
            parts.append(f'<tr><td>конец</td><td>{html.escape(data["last_ts"])}</td>'
                         f'<td>{html.escape(os.path.basename(data["last_file"]))}</td></tr>')
            parts.append('</table></details>')

        # Коды
        if codes:
            sorted_codes = sorted(codes)
            parts.append(f'<details><summary>🔑 Уникальные коды '
                         f'({len(sorted_codes):,})</summary><ul>')
            for c in sorted_codes:
                parts.append(f'<li>{html.escape(c)}</li>')
            parts.append('</ul></details>')

        parts.append('</details>')

    parts.append('</body></html>')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(parts))
    return path, []


# ===========================================================================
# XLSX (openpyxl) — Excel-collapsible через outline levels
# ===========================================================================

def export_xlsx(path, batches):
    """Excel-файл с тремя листами:
      1. «Сводка» — flat: партия, счётчики, итог.
      2. «Файлы» — партия × файл × события.
      3. «Коды» — партия × код.

    На листе «Сводка» используется row outline (excel-grouping) — раскрываемые
    разделы под каждой партией с детализацией её файлов и кодов."""
    openpyxl = _try_import_openpyxl()
    if openpyxl is None:
        raise RuntimeError(
            "Для экспорта в Excel нужна библиотека openpyxl.\n"
            "Установи её: pip install openpyxl"
        )

    from openpyxl.styles import Font, PatternFill, Alignment
    # SGTIN коды содержат GS/FS/RS/US (\x1c-\x1f) как разделители полей
    # DataMatrix — Excel запрещает control chars в ячейках. Чистим их
    # ПЕРЕД записью, иначе openpyxl бросит «cannot be used in worksheets».
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def _safe(val):
        """Удаляет недопустимые в xlsx ASCII control chars из строк."""
        if isinstance(val, str):
            return ILLEGAL_CHARACTERS_RE.sub('', val)
        return val

    wb = openpyxl.Workbook()

    # ---- Лист 1: Сводка с group-collapsible детализацией ----
    ws = wb.active
    ws.title = "Сводка"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2A82DA", end_color="2A82DA",
                              fill_type="solid")
    batch_font = Font(bold=True, color="2A82DA")
    detail_font = Font(italic=True, color="666666")

    headers = ['Партия'] + [label for _, label in _COUNTER_LABELS] + [
        'Всего', 'Файлов', 'Уникальных кодов',
        'Первое время', 'Первый файл',
        'Последнее время', 'Последний файл']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    for bid, data in sorted(batches.items(), key=_sort_key):
        counters = data['counters']
        total_events = sum(counters.values())
        files = data['files']
        codes = data['codes']

        # Главная строка партии — уровень 0
        main_row = ['Вне партии' if bid == NO_BATCH else str(bid)]
        main_row += [counters[k] for k, _ in _COUNTER_LABELS]
        main_row += [
            total_events,
            len(files),
            len(codes),
            data.get('first_ts', ''),
            os.path.basename(data.get('first_file', '')),
            data.get('last_ts', ''),
            os.path.basename(data.get('last_file', '')),
        ]
        ws.append(main_row)
        ws.cell(row=row_idx, column=1).font = batch_font
        row_idx += 1

        # Под-строки для файлов и кодов — outline level 1, hidden=True по
        # умолчанию (свёрнуто). Excel рисует кнопку «+» слева от каждой
        # группы. Юзер кликом раскрывает.
        group_start = row_idx
        for fp in sorted(files.keys()):
            ws.append([f'  📂 {os.path.basename(fp)}',
                       '', '', '', '', '', '',
                       files[fp], '', '', '', '', '', ''])
            ws.cell(row=row_idx, column=1).font = detail_font
            ws.row_dimensions[row_idx].outline_level = 1
            ws.row_dimensions[row_idx].hidden = True
            row_idx += 1
        # Коды — пишем компактно, по 4 на строке. _safe() убирает GS/FS из
        # DataMatrix, которые Excel не разрешает.
        if codes:
            sorted_codes = sorted(codes)
            chunk = 4
            for i in range(0, len(sorted_codes), chunk):
                cell_text = '  🔑 ' + ' / '.join(sorted_codes[i:i + chunk])
                ws.append([_safe(cell_text)] + [''] * (len(headers) - 1))
                ws.cell(row=row_idx, column=1).font = detail_font
                ws.row_dimensions[row_idx].outline_level = 1
                ws.row_dimensions[row_idx].hidden = True
                row_idx += 1
        if group_start < row_idx:
            # Включаем outline — Excel покажет кнопку «+/-» слева
            ws.sheet_properties.outlinePr.summaryBelow = False

    # Ширина столбцов
    ws.column_dimensions['A'].width = 50
    for col_letter in 'BCDEFGHIJKLMN':
        ws.column_dimensions[col_letter].width = 14
    ws.freeze_panes = 'B2'

    # ---- Лист 2: Файлы ----
    ws_files = wb.create_sheet("Файлы")
    ws_files.append(['Партия', 'Файл', 'Событий'])
    for col in range(1, 4):
        ws_files.cell(row=1, column=col).font = header_font
        ws_files.cell(row=1, column=col).fill = header_fill
    for bid, data in sorted(batches.items(), key=_sort_key):
        label = 'Вне партии' if bid == NO_BATCH else str(bid)
        for fp in sorted(data['files'].keys()):
            ws_files.append([label, os.path.basename(fp),
                             data['files'][fp]])
    ws_files.column_dimensions['A'].width = 18
    ws_files.column_dimensions['B'].width = 40
    ws_files.column_dimensions['C'].width = 12
    ws_files.freeze_panes = 'A2'

    # ---- Лист 3: Коды ----
    ws_codes = wb.create_sheet("Коды")
    ws_codes.append(['Партия', 'Код'])
    for col in range(1, 3):
        ws_codes.cell(row=1, column=col).font = header_font
        ws_codes.cell(row=1, column=col).fill = header_fill
    for bid, data in sorted(batches.items(), key=_sort_key):
        label = 'Вне партии' if bid == NO_BATCH else str(bid)
        for c in sorted(data['codes']):
            # _safe() убирает GS/FS из DataMatrix-разделителей кода —
            # без этого Excel ругается «cannot be used in worksheets».
            ws_codes.append([label, _safe(c)])
    ws_codes.column_dimensions['A'].width = 18
    ws_codes.column_dimensions['B'].width = 50
    ws_codes.freeze_panes = 'A2'

    wb.save(path)
    return path, []
